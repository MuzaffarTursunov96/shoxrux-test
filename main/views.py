from django.shortcuts import render,redirect
from django.core.paginator import Paginator
from django.http import JsonResponse
from .models import Quetion
import openpyxl
import os
import random

def index(request):
  return render(request,'base.html')

def test(request,pk):
  page = pk
  if not page:
    page=1
  items = list(Quetion.objects.all())
  questions = random.sample(items, 50)
  # paginator = Paginator(questions,5)
  # questions = paginator.page(page)
  # questions=Quetion.objects.all()
  context ={
    'questions':questions,
    'question_count':50
  }
  return render(request,'index.html',context)
  

def answer(request,pk):
  answer =request.GET.get('answer')
  question =Quetion.objects.get(pk=int(pk))
  if question:
    if question.answer == answer:
      right =True
      answer=answer
    else:
      right=False
      answer=question.answer
    return JsonResponse({'msg':right,'answer':answer})
  else:
    return JsonResponse({'msg':'Nothing found','answer':None})

def xsl(request):
  absolute_path = os.path.abspath(__file__)
  dataframe = openpyxl.load_workbook(f"G:\\rabochiy stol\\shoxruh\\main\\savol.xlsx")
  dataframe1 = dataframe.active
  q=Quetion.objects.all()
  q.delete()
  for row in range(0, dataframe1.max_row):
      project =[]
      question=Quetion()
      for col in dataframe1.iter_cols(1, dataframe1.max_column):
        project.append(col[row].value)
        print(col[row].value)
      print(project)
      
      print('#'*20)
      print(project)
      if project[5] ==None or project[1]==None or project[2]==None or project[3]==None or project[6]==None:
        continue
      question.title =project[1]
      question.a =project[2]
      question.b =project[3]
      question.c =project[5]
      question.d =project[6]
      question.answer ='a'
      question.save()

  #     print(len(project))
  return redirect('index')

def do_rand(request):
  questions =Quetion.objects.all()

  mylist = ["a", "b", "c","d"]

  # print(random.choice(mylist))
  for question in questions:
    rannst=random.choice(mylist)
    if rannst =='b':
      a=question.a
      b=question.b
      question.a=b
      question.b=a
      question.answer ='b'
    elif rannst =='c':
      a=question.a
      c=question.c
      question.a=c
      question.c=a
      question.answer ='c'
    elif rannst =='d':
      a=question.a
      d=question.d
      question.a=d
      question.d=a
      question.answer ='d'
    question.save()
  return redirect('index')